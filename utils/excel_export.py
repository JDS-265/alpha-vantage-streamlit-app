from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# ============================================================
# EXCEL FORMATTING HELPERS
# ============================================================

def sanitize_sheet_name(sheet_name: str) -> str:
    """
    Limpa nomes de folhas para respeitar as regras do Excel.

    Regras principais:
    - máximo 31 caracteres;
    - não pode conter caracteres inválidos: []:*?/\\
    """

    invalid_characters = ["[", "]", ":", "*", "?", "/", "\\"]

    clean_name = str(sheet_name)

    for character in invalid_characters:
        clean_name = clean_name.replace(character, "-")

    clean_name = clean_name.strip()

    if clean_name == "":
        clean_name = "Sheet"

    return clean_name[:31]


def make_excel_safe_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepara um DataFrame para exportação Excel.

    Resolve problemas comuns:
    - colunas não-string;
    - datas timezone-aware;
    - valores pd.NA;
    - objetos mistos.
    """

    if df is None:
        return pd.DataFrame()

    safe_df = df.copy()

    safe_df.columns = [
        str(column)
        for column in safe_df.columns
    ]

    for column in safe_df.columns:
        if pd.api.types.is_datetime64_any_dtype(safe_df[column]):
            safe_df[column] = pd.to_datetime(
                safe_df[column],
                errors="coerce"
            )

            try:
                safe_df[column] = safe_df[column].dt.tz_localize(None)
            except TypeError:
                pass

        elif safe_df[column].dtype == "object":
            safe_df[column] = safe_df[column].apply(
                lambda value: "" if pd.isna(value) else value
            )

    return safe_df


def apply_basic_sheet_formatting(
    workbook_path: Path,
    freeze_panes: str = "A2"
) -> None:
    """
    Aplica formatação profissional simples a todas as folhas do Excel.

    Inclui:
    - header destacado;
    - filtros;
    - freeze panes;
    - ajuste de larguras;
    - bordas leves;
    - alinhamento.
    """

    workbook = load_workbook(workbook_path)

    header_fill = PatternFill(
        start_color="1F4E78",
        end_color="1F4E78",
        fill_type="solid"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    title_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )

    title_font = Font(
        color="1F4E78",
        bold=True,
        size=14
    )

    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3")
    )

    for worksheet in workbook.worksheets:
        max_row = worksheet.max_row
        max_column = worksheet.max_column

        if max_row == 0 or max_column == 0:
            continue

        worksheet.freeze_panes = freeze_panes

        if max_row >= 1 and max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = thin_border

        for row in worksheet.iter_rows(
            min_row=2,
            max_row=max_row,
            max_col=max_column
        ):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=False
                )

                if isinstance(cell.value, float):
                    cell.number_format = "0.00"

                if isinstance(cell.value, int):
                    cell.number_format = "#,##0"

                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd"

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(column_cells[0].column)

            max_length = 0

            for cell in column_cells:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            adjusted_width = min(
                max(max_length + 2, 10),
                35
            )

            worksheet.column_dimensions[column_letter].width = adjusted_width

        worksheet.row_dimensions[1].height = 24

        # Aplicar estilo especial a folhas de README/App Summary
        if worksheet.title in ["README", "App Summary"]:
            for cell in worksheet[1]:
                cell.fill = title_fill
                cell.font = title_font
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True
                )

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        vertical="top",
                        wrap_text=True
                    )

    workbook.save(workbook_path)


# ============================================================
# CORE EXPORT FUNCTIONS
# ============================================================

def write_dataframe_to_excel(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet_name: str,
    include_index: bool = False
) -> None:
    """
    Escreve um DataFrame numa folha Excel.

    Se o DataFrame estiver vazio, escreve uma folha com uma mensagem simples.
    """

    clean_sheet_name = sanitize_sheet_name(sheet_name)

    if df is None or df.empty:
        empty_df = pd.DataFrame({
            "Message": [
                "No data available for this section."
            ]
        })

        empty_df.to_excel(
            writer,
            sheet_name=clean_sheet_name,
            index=False
        )

        return

    safe_df = make_excel_safe_dataframe(df)

    safe_df.to_excel(
        writer,
        sheet_name=clean_sheet_name,
        index=include_index
    )


def create_readme_sheet(
    app_version: str,
    generated_at: datetime,
    selected_stocks: list[str],
    notes: list[str] | None = None
) -> pd.DataFrame:
    """
    Cria a folha README do relatório Excel.
    """

    if notes is None:
        notes = []

    rows = [
        {
            "Section": "Report",
            "Description": "Alpha Vantage Financial Dashboard Excel Export"
        },
        {
            "Section": "App Version",
            "Description": app_version
        },
        {
            "Section": "Generated At",
            "Description": generated_at.strftime("%Y-%m-%d %H:%M:%S")
        },
        {
            "Section": "Selected Stocks",
            "Description": ", ".join(selected_stocks) if selected_stocks else "-"
        },
        {
            "Section": "Purpose",
            "Description": (
                "This workbook consolidates data loaded in the Streamlit app, "
                "including stocks, fundamentals, commodities, FX, crypto and macro indicators."
            )
        },
        {
            "Section": "Important Limitation",
            "Description": (
                "Historical data and calculated metrics are not forecasts and do not represent investment advice."
            )
        },
        {
            "Section": "Data Source",
            "Description": "Alpha Vantage API and local cache generated by the app."
        },
    ]

    for note in notes:
        rows.append({
            "Section": "Note",
            "Description": note
        })

    return pd.DataFrame(rows)


def create_app_summary_sheet(
    app_version: str,
    selected_stocks: list[str],
    export_sections: dict
) -> pd.DataFrame:
    """
    Cria uma folha de resumo da app e das secções exportadas.
    """

    rows = [
        {
            "Item": "App Version",
            "Value": app_version
        },
        {
            "Item": "Selected Stocks",
            "Value": ", ".join(selected_stocks) if selected_stocks else "-"
        },
        {
            "Item": "Exported Sections",
            "Value": ", ".join([
                section
                for section, is_available in export_sections.items()
                if is_available
            ])
        }
    ]

    for section, is_available in export_sections.items():
        rows.append({
            "Item": f"{section} Available",
            "Value": "Yes" if is_available else "No"
        })

    return pd.DataFrame(rows)


def export_alpha_vantage_report(
    output_path: Path,
    app_version: str,
    selected_stocks: list[str],
    dataframes: dict[str, pd.DataFrame],
    notes: list[str] | None = None
) -> Path:
    """
    Exporta um relatório Excel completo da app Alpha Vantage.

    Parâmetros:
    - output_path: caminho final do ficheiro Excel;
    - app_version: versão atual da app;
    - selected_stocks: lista de tickers;
    - dataframes: dicionário com nome_da_folha -> DataFrame;
    - notes: notas opcionais para a folha README.

    Exemplo de dataframes:
    {
        "Stock Summary": stock_summary,
        "Stock Prices": stock_prices,
        "Fundamentals": fundamentals,
        "Macro Summary": macro_summary
    }
    """

    output_path = Path(output_path)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    generated_at = datetime.now()

    export_sections = {
        sheet_name: (
            dataframe is not None
            and isinstance(dataframe, pd.DataFrame)
            and not dataframe.empty
        )
        for sheet_name, dataframe in dataframes.items()
    }

    readme_df = create_readme_sheet(
        app_version=app_version,
        generated_at=generated_at,
        selected_stocks=selected_stocks,
        notes=notes
    )

    app_summary_df = create_app_summary_sheet(
        app_version=app_version,
        selected_stocks=selected_stocks,
        export_sections=export_sections
    )

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl"
    ) as writer:

        write_dataframe_to_excel(
            writer=writer,
            df=readme_df,
            sheet_name="README",
            include_index=False
        )

        write_dataframe_to_excel(
            writer=writer,
            df=app_summary_df,
            sheet_name="App Summary",
            include_index=False
        )

        for sheet_name, dataframe in dataframes.items():
            write_dataframe_to_excel(
                writer=writer,
                df=dataframe,
                sheet_name=sheet_name,
                include_index=False
            )

    apply_basic_sheet_formatting(
        workbook_path=output_path,
        freeze_panes="A2"
    )

    return output_path


# ============================================================
# UTILITY FUNCTION FOR FILE NAMING
# ============================================================

def generate_export_filename(
    prefix: str = "alpha_vantage_report"
) -> str:
    """
    Gera um nome de ficheiro Excel com timestamp.
    """

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    return f"{prefix}_{timestamp}.xlsx"